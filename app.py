from flask import Flask, request, render_template, send_file, redirect, url_for, session, jsonify, make_response
from flask_cors import CORS
import os
import pandas as pd
from itertools import combinations
import tempfile
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import quote
import logging
import mimetypes
import re
from openpyxl.utils import get_column_letter
import traceback  # トレースバック情報の取得用
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# CSPを設定するデコレータ
def add_csp_headers(response):
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' chrome-extension:; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: chrome-extension:; "
        "connect-src 'self' chrome-extension:; "
        "frame-src 'self' chrome-extension:;"
    )
    response.headers['Content-Security-Policy'] = csp
    return response

# ロギングの設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Flask アプリケーションの設定
app = Flask(__name__)
CORS(app, resources={
    r"/*": {
        "origins": ["http://localhost:5000", "http://127.0.0.1:5000", "chrome-extension://*"],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"],
        "supports_credentials": True
    }
}) # credentials サポートを追加

# セッションの設定
app.config.update(
    SECRET_KEY=os.urandom(24),
    SESSION_COOKIE_SECURE=True,  # 開発環境ではFalse
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax'
)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))

# 一時ディレクトリの設定
UPLOAD_FOLDER = tempfile.mkdtemp()
OUTPUT_FOLDER = tempfile.mkdtemp()
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config.update(
    UPLOAD_FOLDER=UPLOAD_FOLDER,
    OUTPUT_FOLDER=OUTPUT_FOLDER,
    MAX_CONTENT_LENGTH=32 * 1024 * 1024,  # 32MB
    TEMPLATES_AUTO_RELOAD=True
)

ALLOWED_EXTENSIONS = {'csv', 'xlsx'}
executor = ThreadPoolExecutor(max_workers=4)

# Flaskアプリケーションで、より緩和されたCSP設定を適用
@app.after_request
def add_security_headers(response):
    # CSPヘッダー
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval' chrome-extension:; style-src 'self' 'unsafe-inline'; connect-src 'self' chrome-extension:;"
    
    # その他のセキュリティヘッダー
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    return response

# ユーティリティ関数
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def sanitize_filename(filename):
    """安全でない文字を削除しつつ、日本語や特定文字を許容"""
    invalid_chars = '<>:"/\\|?*\0'
    clean_name = ''.join(c for c in filename if c not in invalid_chars)
    clean_name = clean_name.strip()
    return clean_name

def clean_campaign_name(campaign_name):
    """キャンペーン名から先頭の数字と全角/半角スラッシュを削除"""
    if pd.isna(campaign_name):
        return campaign_name
    
    campaign_name = str(campaign_name)
    patterns = [
        r'^\d+[/／]',
        r'^\d+\s*[/／]',
        r'^[\d\s/／]+',
        r'^[/／]+',
        r'^\d+[/／]\d+[/／]',
    ]
    
    cleaned = campaign_name
    for pattern in patterns:
        cleaned = re.sub(pattern, '', cleaned)
    
    cleaned = re.sub(r'^[/／]+', '', cleaned)
    cleaned = cleaned.strip()
    
    return cleaned

def find_header_row(filepath, file_ext):
    """データフレームからヘッダー行を特定する（改善版）"""
    # 検索するカラム名のパターン
    id_patterns = ['id', 'ID', '担当者', '見込客']
    campaign_patterns = ['キャンペーン', 'campaign', 'CAMPAIGN']
    
    try:
        # ファイルの読み込み方法を決定
        if file_ext == 'xlsx':
            # Excelファイルの場合、シート名を取得
            xls = pd.ExcelFile(filepath)
            sheet_name = xls.sheet_names[0]  # 最初のシートを使用
            
            # 最初の数行を読み込んでヘッダーを確認
            for i in range(20):  # 最初の20行をチェック
                try:
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=i)
                    if check_columns(df, id_patterns, campaign_patterns):
                        return i, df
                except Exception as e:
                    continue
        else:
            # CSVファイルの場合、エンコーディングを試行
            encodings = ['utf-8', 'shift-jis', 'cp932']
            
            for encoding in encodings:
                try:
                    # 最初の数行を読み込んでヘッダーを確認
                    for i in range(20):  # 最初の20行をチェック
                        try:
                            df = pd.read_csv(filepath, encoding=encoding, header=i)
                            if check_columns(df, id_patterns, campaign_patterns):
                                return i, df
                        except Exception as e:
                            continue
                except Exception as e:
                    continue
                    
        # ヘッダーが見つからない場合
        raise ValueError("必要な列が見つかりませんでした。")
        
    except Exception as e:
        logger.error(f"ファイル読み込みエラー: {str(e)}")
        raise

def check_columns(df, id_patterns, campaign_patterns):
    """データフレームのカラムをチェック"""
    if df.empty or len(df.columns) == 0:
        return False
        
    # カラム名を正規化（空白を削除し、小文字に変換）
    columns = [str(col).strip().lower() for col in df.columns]
    
    # ID列とキャンペーン列の存在確認
    has_id = any(any(pattern.lower() in col for pattern in id_patterns) for col in columns)
    has_campaign = any(any(pattern.lower() in col for pattern in campaign_patterns) for col in columns)
    
    return has_id and has_campaign

def get_column_names(df):
    """必要なカラム名を特定する（改善版）"""
    id_patterns = ['id', 'ID', '担当者', '見込客']
    campaign_patterns = ['キャンペーン', 'campaign', 'CAMPAIGN']
    
    id_column = None
    campaign_column = None
    
    # カラム名を走査
    for col in df.columns:
        col_str = str(col).strip().lower()
        
        # ID列の検索
        if not id_column and any(pattern.lower() in col_str for pattern in id_patterns):
            id_column = col
            
        # キャンペーン列の検索
        if not campaign_column and any(pattern.lower() in col_str for pattern in campaign_patterns):
            campaign_column = col
            
        # 両方見つかった場合は終了
        if id_column and campaign_column:
            break
            
    if not id_column or not campaign_column:
        raise ValueError(f"必要なカラムが見つかりません。\n見つかったカラム: {', '.join(df.columns)}")
        
    return id_column, campaign_column

def process_cooccurrence_file(filepath, original_filename):
    try:
        # ファイルの拡張子を取得
        file_ext = original_filename.rsplit('.', 1)[1].lower()
        
        # ヘッダー行の特定とデータ読み込み
        header_row, df = find_header_row(filepath, file_ext)
        
        # 必要なカラム名を特定
        id_column, campaign_column = get_column_names(df)
        
        # 必要な列のみを抽出して列名を標準化
        df_processed = df[[id_column, campaign_column]].copy()
        df_processed.columns = ["見込客/担当者ID18", "キャンペーン名"]

        # 共起行列の作成
        co_occurrence_counts = {}
        for _, group in df_processed.groupby("見込客/担当者ID18"):
            campaign_names = group["キャンペーン名"].unique()
            if len(campaign_names) >= 2:
                for i in range(len(campaign_names)):
                    for j in range(i + 1, len(campaign_names)):
                        camp1, camp2 = campaign_names[i], campaign_names[j]
                        pair = tuple(sorted([camp1, camp2]))
                        co_occurrence_counts[pair] = co_occurrence_counts.get(pair, 0) + 1

        # ユニークなキャンペーン名を取得
        unique_campaigns = sorted(df_processed["キャンペーン名"].unique())
        
        # 共起行列の作成
        co_occurrence_matrix = pd.DataFrame(0, 
                                         index=unique_campaigns, 
                                         columns=unique_campaigns)
        
        # 共起回数を行列に設定
        for (camp1, camp2), count in co_occurrence_counts.items():
            co_occurrence_matrix.at[camp1, camp2] = count
            co_occurrence_matrix.at[camp2, camp1] = count

        # 合計行を追加（1行のみ）
        sums = co_occurrence_matrix.sum()
        co_occurrence_matrix.loc['合計'] = sums
        
        # 結果ファイルの保存
        base_name = os.path.splitext(original_filename)[0]
        output_filename = f"共起行列_{base_name}.{file_ext}"
        output_filepath = os.path.join(app.config["OUTPUT_FOLDER"], output_filename)
        
        # ファイル形式に応じて保存
        if file_ext == 'xlsx':
            with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                co_occurrence_matrix.to_excel(writer, sheet_name='共起行列')
                
                # シートの取得と整形
                worksheet = writer.sheets['共起行列']
                
                # 列幅の自動調整
                for idx, col in enumerate(co_occurrence_matrix.columns):
                    max_length = max(
                        co_occurrence_matrix[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 2)].width = max_length + 2

                # 合計行のスタイル設定
                last_row = len(co_occurrence_matrix)
                for col in range(1, len(co_occurrence_matrix.columns) + 2):
                    cell = worksheet.cell(row=last_row + 1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        else:
            co_occurrence_matrix.to_csv(output_filepath, encoding='utf-8-sig')

        return output_filename

    except Exception as e:
        logger.error(f"ファイル処理エラー: {str(e)}", exc_info=True)
        raise

def process_campaign_file(filepath, original_filename):
    """キャンペーンファイルの処理（改善版）"""
    try:
        # ファイルの拡張子を取得
        file_ext = original_filename.rsplit('.', 1)[1].lower()
        
        # ファイルの存在確認
        if not os.path.exists(filepath):
            raise ValueError(f"ファイルが見つかりません: {filepath}")
        
        # ファイルサイズの確認
        if os.path.getsize(filepath) == 0:
            raise ValueError("ファイルが空です")
            
        try:
            # ヘッダー行の特定とデータ読み込み
            header_row, df = find_header_row(filepath, file_ext)
            logger.info(f"ヘッダー行を特定: {header_row}行目")
            logger.info(f"検出されたカラム: {', '.join(df.columns)}")
            
        except Exception as e:
            logger.error(f"ヘッダー行の特定に失敗: {str(e)}")
            raise ValueError(f"データ構造を認識できませんでした。\nエラー: {str(e)}")
        
        # 必要なカラム名を特定
        try:
            id_column, campaign_column = get_column_names(df)
            logger.info(f"特定されたカラム - ID: {id_column}, キャンペーン: {campaign_column}")
            
        except Exception as e:
            logger.error(f"カラム名の特定に失敗: {str(e)}")
            raise ValueError(f"必要なカラムが見つかりません。\nエラー: {str(e)}")
        
        # キャンペーン名のクリーニング
        df[campaign_column] = df[campaign_column].apply(clean_campaign_name)
        
        # 出力ファイルの準備
        output_filename = f"cleaned_{sanitize_filename(original_filename)}"
        output_filepath = os.path.join(app.config["OUTPUT_FOLDER"], output_filename)
        
        # ファイルの保存
        if file_ext == 'csv':
            df.to_csv(output_filepath, encoding='utf-8-sig', index=False)
        else:  # xlsx
            df.to_excel(output_filepath, index=False)
            
        logger.info(f"ファイル処理完了: {output_filename}")
        return output_filename

    except Exception as e:
        logger.error(f"ファイル処理エラー: {str(e)}", exc_info=True)
        raise

# ルート定義
@app.route("/")
def index():
    """トップページ（共起行列生成ツール）"""
    return render_template("co_matrix.html")

@app.route("/campaign")
def campaign():
    """キャンペーン名統合ツール"""
    return render_template("campaign_int.html")

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'files[]' not in request.files:
            return jsonify({
                'success': False,
                'error': 'ファイルがありません'
            }), 400
        
        files = request.files.getlist('files[]')
        uploaded_files = []
        
        for file in files:
            if file and allowed_file(file.filename):
                filename = sanitize_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                uploaded_files.append(filename)
        
        if not uploaded_files:
            return jsonify({
                'success': False,
                'error': '有効なファイルがありません。CSVまたはXLSXファイルを選択してください。'
            }), 400
        
        return jsonify({
            'success': True,
            'files': uploaded_files
        })

    except Exception as e:
        logger.error(f"アップロードエラー: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'ファイルアップロード中にエラーが発生しました: {str(e)}'
        }), 500

@app.route('/process_cooccurrence', methods=['POST'])
def process_cooccurrence_files():
    """共起行列生成処理のエンドポイント"""
    try:
        # リクエストのバリデーション
        if not request.is_json:
            return jsonify({
                'success': False,
                'error': '無効なリクエスト形式です'
            }), 400

        data = request.get_json()
        if not data or 'files' not in data:
            return jsonify({
                'success': False,
                'error': 'ファイル情報が含まれていません'
            }), 400

        filenames = data['files']
        if not filenames:
            return jsonify({
                'success': False,
                'error': '処理するファイルがありません'
            }), 400

        # 処理結果の追跡
        output_files = []
        errors = []

        # 各ファイルを処理
        for filename in filenames:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                if not os.path.exists(filepath):
                    errors.append(f"ファイルが見つかりません: {filename}")
                    continue

                result = process_cooccurrence_file(filepath, filename)
                if result:
                    output_files.append(result)
                    logger.info(f"処理成功: {filename} -> {result}")

            except Exception as e:
                error_msg = f"{filename}: {str(e)}"
                logger.error(f"ファイル処理エラー: {error_msg}", exc_info=True)
                errors.append(error_msg)

        # 処理結果の確認と応答の生成
        if not output_files and errors:
            return jsonify({
                'success': False,
                'error': '全てのファイルの処理に失敗しました',
                'errors': errors
            }), 400

        # セッションに結果を保存
        session['output_files'] = output_files
        
        response_data = {
            'success': True,
            'redirect': url_for('complete_cooccurrence'),
            'processed_files': output_files
        }
        if errors:
            response_data['warnings'] = errors

        logger.info(f"処理完了 - レスポンス: {response_data}")
        return jsonify(response_data)

    except Exception as e:
        logger.error(f"予期せぬエラー: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'予期せぬエラーが発生しました: {str(e)}'
        }), 500

def process_cooccurrence_file(filepath, original_filename):
    try:
        # [前略: ファイル読み込みまでの部分は変更なし]

        # 共起行列の作成
        co_occurrence_counts = {}
        for _, group in data.groupby("見込客/担当者ID18"):
            campaign_names = group["キャンペーン名"].unique()
            if len(campaign_names) >= 2:
                for i in range(len(campaign_names)):
                    for j in range(i + 1, len(campaign_names)):
                        camp1, camp2 = campaign_names[i], campaign_names[j]
                        pair = tuple(sorted([camp1, camp2]))
                        co_occurrence_counts[pair] = co_occurrence_counts.get(pair, 0) + 1

        # ユニークなキャンペーン名を取得
        unique_campaigns = sorted(data["キャンペーン名"].unique())
        
        # 共起行列の作成
        co_occurrence_matrix = pd.DataFrame(0, 
                                         index=unique_campaigns, 
                                         columns=unique_campaigns)
        
        # 共起回数を行列に設定
        for (camp1, camp2), count in co_occurrence_counts.items():
            co_occurrence_matrix.at[camp1, camp2] = count
            co_occurrence_matrix.at[camp2, camp1] = count

        # 合計行を追加（1行のみ）
        sums = co_occurrence_matrix.sum()
        co_occurrence_matrix.loc['合計'] = sums
        
        # 結果ファイルの保存
        base_name = os.path.splitext(original_filename)[0]
        output_filename = f"共起行列_{base_name}.{file_ext}"
        output_filepath = os.path.join(app.config["OUTPUT_FOLDER"], output_filename)
        
        # ファイル形式に応じて保存
        if file_ext == 'xlsx':
            with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                co_occurrence_matrix.to_excel(writer, sheet_name='共起行列')
                
                # シートの取得と整形
                worksheet = writer.sheets['共起行列']
                
                # 列幅の自動調整
                for idx, col in enumerate(co_occurrence_matrix.columns):
                    max_length = max(
                        co_occurrence_matrix[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 2)].width = max_length + 2

                # 合計行のスタイル設定
                last_row = len(co_occurrence_matrix)
                for col in range(1, len(co_occurrence_matrix.columns) + 2):
                    cell = worksheet.cell(row=last_row + 1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        else:
            co_occurrence_matrix.to_csv(output_filepath, encoding='utf-8-sig')

        return output_filename

    except Exception as e:
        logger.error(f"ファイル処理エラー: {str(e)}", exc_info=True)
        raise
    
    
@app.route('/process_campaign', methods=['POST'])
def process_campaign_files():
    """キャンペーン名クリーニング処理のエンドポイント"""
    try:
        # リクエストのログ出力
        logger.info(f"Received request: {request.get_json() if request.is_json else 'Not JSON'}")

        if not request.is_json:
            return jsonify({
                'success': False,
                'error': '無効なリクエスト形式です'
            }), 400

        data = request.get_json()
        if not data or 'files' not in data:
            return jsonify({
                'success': False,
                'error': 'ファイル情報が含まれていません'
            }), 400

        filenames = data['files']
        if not filenames:
            return jsonify({
                'success': False,
                'error': '処理するファイルがありません'
            }), 400

        # 処理結果の追跡
        output_files = []
        errors = []

        # 各ファイルを処理
        for filename in filenames:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                # ファイルの存在確認
                if not os.path.exists(filepath):
                    error_msg = f"ファイルが見つかりません: {filename}"
                    logger.error(error_msg)
                    errors.append(error_msg)
                    continue

                # ファイルの処理
                logger.info(f"Processing file: {filename}")
                result = process_campaign_file(filepath, filename)
                
                if result:
                    output_files.append(result)
                    logger.info(f"Successfully processed: {filename} -> {result}")

            except Exception as e:
                error_msg = f"{filename}: {str(e)}"
                logger.error(f"File processing error: {error_msg}")
                logger.error(traceback.format_exc())  # スタックトレースを出力
                errors.append(error_msg)

        # 処理結果の確認
        if not output_files and errors:
            return jsonify({
                'success': False,
                'error': '全てのファイルの処理に失敗しました',
                'errors': errors
            }), 400

        # セッションに結果を保存
        session['output_files'] = output_files
        
        response_data = {
            'success': True,
            'redirect': url_for('complete_campaign'),
            'processed_files': output_files
        }
        if errors:
            response_data['warnings'] = errors

        logger.info(f"Processing complete - Response: {response_data}")
        return jsonify(response_data)

    except Exception as e:
        error_msg = f"予期せぬエラー: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())  # スタックトレースを出力
        return jsonify({
            'success': False,
            'error': error_msg
        }), 500

@app.route("/complete_cooccurrence")
def complete_cooccurrence():
    """共起行列生成完了ページ"""
    output_files = session.get('output_files', [])
    return render_template(
        "complete_top.html", 
        output_files=output_files,
        file_count=len(output_files)
    )

@app.route("/complete_campaign")
def complete_campaign():
    """キャンペーン名クリーニング完了ページ"""
    output_files = session.get('output_files', [])
    return render_template("complete_second.html", output_files=output_files)

@app.route("/download/<path:filename>")
def download_file(filename):
    try:
        filepath = os.path.join(app.config["OUTPUT_FOLDER"], filename)
        if not os.path.exists(filepath):
            return "ファイルが見つかりません", 404

        # ファイルの種類に応じてMIMEタイプを設定
        file_ext = filename.rsplit('.', 1)[1].lower()
        if file_ext == 'xlsx':
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mimetype = 'text/csv;charset=utf-8'

        response = send_file(
            filepath,
            as_attachment=True,
            mimetype=mimetype,
            download_name=filename
        )
        response.headers["Content-Disposition"] = \
            f"attachment; filename*=UTF-8''{quote(filename)}"
        
        return response

    except Exception as e:
        logger.error(f"ダウンロードエラー: {str(e)}", exc_info=True)
        return jsonify({'error': 'ダウンロードに失敗しました'}), 500

@app.errorhandler(500)
def internal_server_error(e):
    logger.error(f"500 error: {str(e)}")
    logger.error(traceback.format_exc())
    return jsonify({
        'success': False,
        'error': 'サーバー内部エラーが発生しました'
    }), 500

@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Unhandled exception: {str(e)}")
    logger.error(traceback.format_exc())
    return jsonify({
        'success': False,
        'error': 'サーバーエラーが発生しました'
    }), 500


if __name__ == "__main__":
    app.run(debug=False)